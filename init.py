import glob
import os
import cv2
import shutil
from shutil import copyfile
from pdf2jpg import pdf2jpg
import numpy as np
from numpy import array
import xlsxwriter
#import openpyxl
from openpyxl import Workbook, load_workbook
import pytesseract
from imutils.object_detection import non_max_suppression
import time
from time import localtime, strftime, sleep
import tkinter as tk
from tkinter import Tk,ttk,filedialog
from PIL import Image, ImageTk
from PyPDF2 import PdfFileReader
import math
from math import floor, ceil
#from lines_after import *
import ezdxf
#from linetag import *
import logging
#import io
import xlrd
import csv
import System_details as sd
import datetime
from symbol_text_wrapper import detect_text_symbol
from lines_and_text import lines_and_text
from symbols import *
import argparse
from symbols import *
import sys
import linecache

 
def count_PDF_pages(inputpath):
    """
    This function returns the number of pages across all PDFs and the dimension of each page.
    Arguments:
        inputpath {str} -- Path to the input PDFs on disk
    
    Returns:
        (int, {str:(int, int)}) -- A tuple containing the number of pages, and a dictionary containing the width and height of all pages
    """    
    try:
        ret = {}
        total_PDF_pages = 0
        for file in glob.glob(inputpath + "\\" + "*.PDF"):
            pdf = PdfFileReader(open(file,'rb'))
            total_PDF_pages = total_PDF_pages + pdf.getNumPages()
            for i in range(pdf.getNumPages()):
                page = pdf.getPage(i)
                w = (float(page.mediaBox.getWidth())/72)
                h = (float(page.mediaBox.getHeight())/72)
                ret[str(i) + "_" + os.path.basename(file) + '.jpg'] = (min(h, w), max(h, w))
           
        #error_present = 0;
        return total_PDF_pages, ret
    except:
        error_present = 1
        total_PDF_pages = -1
        logging.error(' Error in "count_PDF_pages" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise
    

    

def pdf_to_jpg(inputpath, imagepath):
    """
    This function is used to convert PDF drawings into JPGs

    Arguments:
        inputpath {[type]} -- [description]
        imagepath {[type]} -- [description]
    """    
    try:
        global step_now
        for file in glob.glob(inputpath + "\\" + "*.pdf"):
            logging.info(' Converting ' + file + " to JPG")  
            pdf2jpg.convert_pdf2jpg(file, imagepath, dpi=300, pages="ALL")
            if not __debug__:
                progress_bar_increment()   
                label_1.configure(text = "Converting " + os.path.basename(file) + " to JPG")
                label_1.update()
    except:
        #global error_present
        error_present = 1
        logging.error(' Error in "pdf_to_jpg" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise


def tiff_png_jpg_to_jpg(inputpath, imagepath):
    """
    This function is used to convert PNG,JPG and TIFF drawings into JPGs and transfer them to imagepath
    
    Arguments:
        inputpath {str} -- Path to the directory containing the input images
        imagepath {[type]} -- Path to the temporary directory containing newly converted images
    """    
    try:
        global step_now 
        png_files = glob.glob(inputpath + "\\" + "*.png")
        tiff_files = glob.glob(inputpath + "\\" + "*.tiff") 
        tif_files = glob.glob(inputpath + "\\" + "*.tif")
        jpg_files = glob.glob(inputpath + "\\" + "*.jpg")
        imagefiles = png_files + tiff_files + tif_files + jpg_files
        for file in imagefiles:
            logging.info(' Converting ' + file + " to JPG")
            img = Image.open(file)
            file_name = file.replace(inputpath + "\\","")
            os.makedirs(imagepath + "\\" + file_name)
            img.save(imagepath + "\\" + file_name + "\\" + "0_" + file_name + ".jpg")
            if not __debug__:
                progress_bar_increment()
                label_1.configure(text = "Converting " + os.path.basename(file) + " to JPG")
                label_1.update()
    except:
        #global error_present
        error_present = 1
        logging.error(' Error in "tiff_png_jpg_to_jpg" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise

def output_files_and_folder_creator(name, outputpath):
    """
    This function creates the output folder for a specific page of engineering drawing PDF
    and also creates the "INSTRUMENTS.xlsx" and "LINES.xlsx" Excel workbooks along with
    the "output_drawing.jpg" in the folder
    
    Arguments:
        name {str} -- Full path to the input file
        outputpath {str} -- Directory where the output files and folders will be placed
    
    Returns:
        (str, str, str, str) -- Page name, Output drawing path, Instrument workbook path, Piping workbook path 
    """    
    try:

        # Creating the folder name with the format PAGE-pagenumber_inputfilename
        temp_name_1 = os.path.basename(name)
        temp_name_1 = temp_name_1.replace(".jpg","")
        split_words = temp_name_1.split("_")
        page_no = split_words[0]
        page_no = int(page_no) + 1
        temp_name_1 = temp_name_1[1:]
        page_name = "PAGE-" + str(page_no) + temp_name_1
        # Removing the folder of the same name if it exists and creates a new one, otherwise creates a folder with the name format specified
        if os.path.isdir(outputpath + "\\" + page_name):
            shutil.rmtree(outputpath + "\\" + page_name)
            os.makedirs(outputpath + "\\" + page_name)
        else:
            os.makedirs(outputpath + "\\" + page_name)
        # Creating a copy of the drawing concerned in the output folder for the purpose of highlighting the instruments and lines" 
        copyfile(name, outputpath + "\\" + page_name + "\\" + "output_drawing.jpg")
        # Creating a copy of the drawing concerned in the program folder for the purpose of removing the instruments once they are detected" 
        copyfile(name, "cropc.jpg")
        output_drawing = outputpath + "\\" + page_name + "\\" + "output_drawing.jpg"
        # Creating Excel sheet for instrument data
        workbook_instrument = outputpath + "\\" + page_name + "\\" + "COMPONENTS_DATA.xlsx"    
        workbook = xlsxwriter.Workbook(workbook_instrument)    
        worksheet = workbook.add_worksheet()
        cell_format = workbook.add_format({'bold': True, 'font_color': 'red'})
        worksheet.write('A1', 'Document Name' , cell_format)
        worksheet.write('B1', 'Component Name', cell_format)
        worksheet.write('C1', 'x Coordinate', cell_format)
        worksheet.write('D1', 'y Coordinate', cell_format)
        worksheet.write('E1', 'Component Tag', cell_format)
        workbook.close()
        # Creating an Excel sheet for piping (line) data
        workbook_name = outputpath + "\\" + page_name + "\\" + "PIPING_DATA.xlsx"       
        workbook = xlsxwriter.Workbook(workbook_name)    
        worksheet = workbook.add_worksheet()
        workbook.close()
        wb = load_workbook(workbook_name)
        dest_filename = workbook_name
        sheet_get= wb.worksheets[0]
        cell_insert = 'A1'
        sheet_get [str(cell_insert)] = "x1"
        cell_insert = 'B1'
        sheet_get [str(cell_insert)] = "y1"
        cell_insert = 'C1'
        sheet_get [str(cell_insert)] = "x2"
        cell_insert = 'D1'
        sheet_get [str(cell_insert)] = "y2"    
        cell_insert = 'E1'
        sheet_get [str(cell_insert)] = "Piping Tag"    
        wb.save(filename = dest_filename)
        
        return page_name, output_drawing, workbook_instrument, workbook_name
    except:
        error_present = 1
        logging.error(' Error in "output_files_and_folder" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise


def get_instrument_and_line_data(imagepath, outputpath, templatepath, pagedims):
    """
    Given paths to the input image directory, the output directory and the template directory, this function will perform symbol, text and line detection on the drawings and will put the outputs in Excel sheets, CSV files and AutoCAD DXF files.
    
    Arguments:
        imagepath {str} -- Path to the directory containing input images
        outputpath {str} -- Path to the directory where the outputs will be stored
        templatepath {str} -- Path to the directory containing instrument templates
        pagedims {{str:(double, double)}} -- Dictionary containing the dimensions of each page in inches
    """    
    try:
        const_EAST_SIZE_FACTOR = 32

        #Getting list of folders in imagepath
        train2 = os.listdir(imagepath)
        
        #Accessing each of these folders
        for file in train2:
            #Accessing every file in the folder inside the imagepath

            for name in glob.glob(imagepath + "\\" + file + "\\" + "*.jpg"):
                #Current progress of the progress bar
                global step_now
               
                instrument_excel_row = 2
                
                dwg = ezdxf.new('R2010', setup=True)
                modelspace = dwg.modelspace()
                #Getting list of template folders inside templatepath
                templates_list = os.listdir(templatepath)

                logging.info(' Creating outout folder for ' + file)
                page_name, output_drawing, workbook_instrument, workbook_name = output_files_and_folder_creator(name, outputpath)
                #Accessing each of these folders
                img = cv2.imread(output_drawing)

                width = 9930
                height = 7017
                dim = (width, height)

                # resize image
                img = cv2.resize(img, dim, interpolation = cv2.INTER_AREA)

                (w, h) = dim
                newh = ceil(img.shape[0]/const_EAST_SIZE_FACTOR) * const_EAST_SIZE_FACTOR
                neww = ceil(img.shape[1]/const_EAST_SIZE_FACTOR) * const_EAST_SIZE_FACTOR
                print(h, w)
                print(newh, neww)
                # Find the amount to pad the image by on both sides
                dh = int((newh - h)/2)
                dw = int((neww - w)/2)
                padding = (dh, dw)
                # Pad the image, store the new dimensions
                img = cv2.copyMakeBorder(img, dh, dh, dw, dw, cv2.BORDER_CONSTANT, value=[255, 255, 255])
                tested_blocks = []
                cv2.imwrite(output_drawing, img)
                cv2.imwrite('cropc.jpg', img)
                for template_name in templates_list:
                    template_instance = 0
                    if not __debug__:
                        label_1.configure(text = "Detecting " + template_name + " in " + page_name)       #Updating the backend data display
                    logging.info(' Detecting ' + template_name + " in " + page_name)
                    #Accessing every template in the folder inside the templatepath
                    for specific_template in glob.glob(templatepath + "\\" + template_name + "\\" + "*.jpg"):
                        logging.info(' Detecting ' + specific_template)
                        #Getting the threshold for the template
                        try:
                            threshold_file = open(templatepath + "\\" + template_name + "\\" + "Threshold.txt", "r")         
                        except:
                            print(" No threshold file found for " + specific_template)
                            logging.critical(" No threshold file found for " + specific_template + ", skipping")
                            continue
                        threshold = threshold_file.read()
                        threshold = float(threshold)
                        logging.info(' Threshold: ' + str(threshold))
                        instrument_excel_row , template_instance = getboxes(output_drawing, specific_template, file, template_name, threshold, workbook_instrument, instrument_excel_row, template_instance, dwg, modelspace, tested_blocks, padding)    #Calling the function "getboxes" for instrument detection
                        if not __debug__:
                            progress_bar_increment()
                        #logging.info(' Progress bar completion: ' + str(floor(step_now)))
                    if template_instance > 0:
                        logging.info(str(template_instance) + ' symbols detected')
                        logging.info(str(template_instance) + ' symbols detected')

                dwg.saveas(os.path.join(outputpath, page_name + '.dxf'))
                dxfpath = os.path.join(outputpath, page_name + '.dxf')               
                if not __debug__: 
                    label_1.configure(text = "Detecting lines and text in " + page_name )     #Updating the backend data display
                    label_1.update()
                    progress_bar_increment()

                logging.info(' Detecting lines in ' + page_name)
                ltout = lines_and_text("cropc.jpg", page_name, workbook_name, output_drawing , dxfpath, padding)
                logging.info("Finished 'get_instrument_and_line_data' function")
    except:
        error_present = 1
        logging.error(' Error in "get_instrument_and_line_data" function : ')
        error_message = PrintException()
        logging.error(error_message)
        #root.destroy()
        raise

def get_total_templates(templatepath):
    """
    This function returns the total number of templates.

    
    Arguments:
        templatepath {str} -- Path to the directory containing instrument templates
    
    Returns:
        int -- Total number of templates in the directory
    """
    try:
        total_template_number = 0    
        templates_list = os.listdir(templatepath)
        for template_name in templates_list:
            for specific_template in glob.glob(templatepath + "\\" + template_name + "\\" + "*.jpg"):
                total_template_number = total_template_number + 1

        return total_template_number
    except:
        error_present = 1
        logging.error(' Error in "get_total_templates" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise

def start_code(inputpath, templatepath, imagepath, outputpath):
    """
    Main function of BrewCAD.
    
    Arguments:
        inputpath {str} -- Path to the directory containing input PDF/image drawings
        templatepath {str} -- Path to the directory containing instrument templates
        imagepath {str} -- Path to the temp directory for converted images
        outputpath {str} -- Path to the directory where the outputs wll be stored
    
    Returns:
        int -- Conbined number of pages, PDFs, images and templates for use in the progress bar and for storing in the Worley database
    """    
    try:
        if not os.path.exists('log'):
            os.mkdir('log')
        working_dir = os.path.abspath(os.path.dirname(__file__))
        logpath = os.path.join(working_dir, 'log')
        logpath = os.path.join(logpath, strftime('%Y-%m-%d'))
        if not os.path.exists(logpath):
            os.mkdir(logpath)
        logname = strftime('%H%M%S') + '.log'
        logfile = os.path.join(logpath, logname)
        try:
            file = open(logfile, "w") 
        except:
            file = open(logfile, 'a')
        if __debug__:
            level = logging.DEBUG
        else:
            level = logging.INFO
        format = '%(asctime)s (%(funcName)s) %(levelname)s: %(message)s'
        logging.basicConfig(level=level, filename=logfile, filemode='w', format=format, datefmt='%d.%m.%Y %H:%M:%S')
        logging.info('Logging configured')
    except:
        raise
    if __debug__:
        logging.info('Debug mode activated...')
    error_present = 0
    try:
        global step_now



        #Getting total number of PDFs
        total_PDFs = 0
        for file in glob.glob(inputpath + "\\" + "*.PDF"):
            total_PDFs = total_PDFs + 1
        
        #Getting total number of JPGs
        total_JPGs = 0
        for file in glob.glob(inputpath + "\\" + "*.jpg"):
            total_JPGs = total_JPGs + 1
        
        #Getting total number of PNGs
        total_PNGs = 0
        for file in glob.glob(inputpath + "\\" + "*.png"):
            total_PNGs = total_PNGs + 1
        
        #Getting total number of TIFFs
        total_TIFFs = 0
        for file in glob.glob(inputpath + "\\" + "*.tiff"):
            total_TIFFs = total_TIFFs + 1
        for file in glob.glob(inputpath + "\\" + "*.tif"):
            total_TIFFs = total_TIFFs + 1
        
        #Getting total number of pages from all PDFs
        logging.info(' Getting the total number of pages from all PDFs')
        total_PDF_pages, pagedims = count_PDF_pages(inputpath)    
        print(pagedims)
        #Getting total number of templates
        logging.info(' Getting the total number of templates')
        total_template_number = get_total_templates(templatepath)    
        global step1 
        step1 = (100/(total_PDFs + total_JPGs + total_PNGs + total_TIFFs + (total_PDF_pages + total_JPGs + total_PNGs + total_TIFFs)*total_template_number + total_PDF_pages))      #Declaring the step size for the progress bar
        step_now = 0

        global NoOfUnits
        NoOfUnits = total_PDF_pages+total_JPGs + total_PNGs + total_TIFFs
    
        #Converting PDFs to JPGs
        pdf_to_jpg(inputpath, imagepath)

        #Converting PNGs, JPGs and TIFFs to JPGs
        tiff_png_jpg_to_jpg(inputpath, imagepath)
        get_instrument_and_line_data(imagepath, outputpath, templatepath, pagedims)
           
        if not __debug__:
            progress1['value'] = 250
            progress1.update()
        time.sleep(2)
        if not __debug__:
            root.destroy()
        
        logging.info("Finished 'start_code' function")
       
        return NoOfUnits
    except:
        #global error_present
        error_present = 1
        logging.error(' Error in "start_code" function : ')
        error_message = PrintException()
        logging.error(error_message)
        if not __debug__:
            root.destroy()
        raise


def progress_bar_increment():
    """
    Increment the progress bar by [step1]
    """
    global step_now
    global step1
    step_now = step_now + step1
    progress1['value'] = floor(step_now)
    progress1.update()
        

class Root1(tk.Tk):
    """
    Class declaration of Tkinter UI for indicating the task completion of BrewCAD.
    """    
    def __init__(self):
        super(Root1, self).__init__()
        self.title("BrewCAD")
        self.minsize(235, 105)
        
        self.label_3 = tk.ttk.Label(self, text = "Instrument detection  has completed successfully.")
        self.label_3.grid(column = 0, row = 0)
    
        self.button4 = tk.ttk.Button(text = "Open Output Folder", command = self.openfolder)
        self.button4.grid(column = 0, row = 5, pady = 15)
        
        self.button5 = tk.ttk.Button(text = "Exit", command = self.destroy)
        self.button5.grid(column = 0, row = 6)
        
    def openfolder(self):
        global outputpath
        os.startfile(outputpath)
        
class Root2(tk.Tk):
    """
    Class declaration of Tkinter UI for indicating an error while running BrewCAD
    
    """    
    def __init__(self):
        super(Root2, self).__init__()
        self.title("BrewCAD")
        self.minsize(235, 105)
        
        self.label_3 = tk.ttk.Label(self, text = "An error has occurred. Please see the log file for further information.")
        self.label_3.grid(column = 0, row = 0)
    
        self.button4 = tk.ttk.Button(text = "Open Log file", command = self.openlogfile)
        self.button4.grid(column = 0, row = 5, pady = 15)
        
        self.button5 = tk.ttk.Button(text = "Exit", command = self.destroy)
        self.button5.grid(column = 0, row = 6)
        
    def openlogfile(self):
        global outputpath
        os.startfile(outputpath + "//" + "logfile.log")


def fileDialog1():
    """
    Function to take user input for "Input Drawing" folder
    """
    try:
            
        filename1 = tk.filedialog.askdirectory()
        label4 = tk.ttk.Label(root, text = "")
        label4.place(x = 25, y = 65)
        label4.configure(text = filename1)
        global inputpath
        inputpath = filename1
    except:
        global error_present
        error_present = 1
        logging.error(' Error in "fileDialog1" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise
    
def fileDialog2():
    """
    Function to take user input for "Output" folder
    """
    try:
        filename2 = tk.filedialog.askdirectory()
        label5 = tk.ttk.Label(root, text = "")
        label5.place(x = 25, y = 165)
        label5.configure(text = filename2)
        global outputpath
        outputpath = filename2
    except:
        global error_present
        error_present = 1
        logging.error(' Error in "fileDialog2" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise
    
def fileDialog3():
    """
    Function to take user input for "Template" folder
    """
    try:
        filename3 = tk.filedialog.askdirectory()
        label6 = tk.ttk.Label(root, text = "")
        label6.place(x = 25, y = 265)
        label6.configure(text = filename3)
        global templatepath
        templatepath = filename3
    except:
        global error_present
        error_present = 1
        logging.error('Error in "fileDialog3" function : ')
        error_message = PrintException()
        logging.error(error_message)
        raise


def callback_error(self, *args):
    """
    Catches any exception in TKinter
    """
    message = 'Error :\n\n'
    message += traceback.format_exc()
    logging.error(message)
    raise


if __name__ == '__main__':
    """
    Main function
    """    
    error_present = 0
    if __debug__:
        parser = argparse.ArgumentParser()
        parser.add_argument("inputpath", help="Path to the input directory")
        parser.add_argument("outputpath", help="Path to the output directory")
        parser.add_argument("templatepath", help="Path to the template directory")
        args = parser.parse_args()
        inputpath = args.inputpath
        templatepath = args.templatepath
        outputpath = args.outputpath
    else:
        inputpath = None
        templatepath = None
        outputpath = None

    path = os.getcwd()
    imagepath = path + "\Image"

    workbook_name = 0
    instrument_excel_row = 0
    template_instance = 0
    StartDateTime = datetime.datetime.now()
    TaskName='BrewCAD'
   
    error_present = 0
    try:
        #Cleaning the folder containing JPGs
        
        if os.path.isdir(imagepath):
            shutil.rmtree(imagepath)
            os.makedirs(imagepath)
        else:
            os.makedirs(imagepath)  
        if __debug__:
            NoOfUnits = start_code(inputpath, templatepath, imagepath, outputpath)
        else:
            #Developing the UI for BrewCAD
            try:
                root = tk.Tk()
                root.geometry('690x460')
                root.title("BrewCAD")
                frame1 = tk.Frame(root, width=670, height=440, bd=1, relief = "solid")
                frame1.place(x = 10, y = 10)
                label1 = tk.ttk.Label(root, text = "Select Input Drawings", font = ("bold",10))
                label1.place(x = 25, y = 40)
                button1 = tk.ttk.Button(root, text = "Browse", command = fileDialog1)
                button1.place(x = 205, y = 35)
                label2 = tk.ttk.Label(root, text = "Select Output Folder", font = ("bold",10))
                label2.place(x = 25, y = 140)
                button2 = tk.ttk.Button(root, text = "Browse", command = fileDialog2)
                button2.place(x = 205, y = 135)
                label3 = tk.ttk.Label(root, text = "Select Template Folder", font = ("bold",10))
                label3.place(x = 25, y = 240)
                button3 = tk.ttk.Button(root, text = "Browse", command = fileDialog3)
                button3.place(x = 205, y = 235)
                button4 = tk.ttk.Button(text = "Start", command = lambda : start_code(inputpath, templatepath, imagepath, outputpath))     #this forks to start the main code
                button4.place(x = 205, y = 335)
                load = Image.open("solo.jpg")
                render = ImageTk.PhotoImage(load)
                img = tk.Label(root, image=render)
                img.image = render
                img.place(x=490, y=30)
                label4 = tk.ttk.Label(root, text = "")
                label4.place(x = 25, y = 85)
                label4.configure(text = inputpath)
                label5 = tk.ttk.Label(root, text = "")
                label5.place(x = 25, y = 185)
                label5.configure(text = outputpath)
                label6 = tk.ttk.Label(root, text = "")
                label6.place(x = 25, y = 285)
                label6.configure(text = templatepath)
    
    
                label_1 = tk.ttk.Label(root, text = "")
                label_1.place(x = 25, y = 380)
                progress1 = tk.ttk.Progressbar(root, length=250)
                progress1.place(x = 25, y = 400)
                tk.Tk.report_callback_exception = callback_error 

                root.mainloop()
            except:
                raise
        
    except:
        logging.error(' Error in main code ')
        error_message = PrintException()
        logging.error(error_message)
        print("An error has occured. Please check the log file for further details.")
        error_present = 1
    finally:
        LoggedInUser=sd.get_logged_in_user()
        MachineName=sd.get_machine_name()
        Country, Location, mail = sd.get_mail_country_and_location()
        AutomationType='DataScience'
        EndDateTime=datetime.datetime.now()
        if error_present == 0:
            Status='Completed'
            print(Status)
            #sd.storing_in_db(TaskName,StartDateTime,EndDateTime,Status,total_units)
            sd.storing_in_db(TaskName,StartDateTime,EndDateTime,Status,LoggedInUser,MachineName,Country,Location,AutomationType,NoOfUnits)
            if not __debug__:
                root = Root1()
                root.mainloop()
            
        else:
            Status = 'Failed'
            print(Status)
            sd.storing_in_db(TaskName,StartDateTime,EndDateTime,Status,LoggedInUser,MachineName,Country,Location,AutomationType,NoOfUnits)
            #sd.storing_in_db(TaskName,StartDateTime,EndDateTime,Status,total_units)
            print("An error has occured. Please see the log file for further information.")
            if not __debug__:            
                root = Root2()
                root.mainloop()
            
        try:
            if os.path.exists("cropa.jpg"):
                os.remove("cropa.jpg")
            if os.path.exists("cropb.jpg"):
                os.remove("cropb.jpg")
            if os.path.exists("cropc.jpg"):
                os.remove("cropc.jpg")
            #Cleaning the image folder at the end of the code
            if os.path.isdir(imagepath):
                shutil.rmtree(imagepath)
            logging.shutdown()
        except:
            print("Error occured while cleaning the temporary files")
            logging.shutdown()